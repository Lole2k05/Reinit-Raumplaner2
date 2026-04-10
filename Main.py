import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog, colorchooser
from PIL import Image, ImageTk
import os
import sys
import json
import errno
import subprocess
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path

# --- GIT UPDATE-FUNKTIONEN ---
REPO_URL = "https://github.com/Lole2k05/Reinit-Raumplaner2.git"

def check_and_pull_updates():
    """Prüft auf Updates und pullt diese automatisch"""
    try:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Prüfe, ob wir in einem Git-Repo sind
        result = subprocess.run(
            ["git", "rev-parse", "--git-dir"],
            cwd=current_dir,
            capture_output=True,
            text=True
        )
        
        if result.returncode != 0:
            print("⚠ Kein Git-Repository gefunden. Initialisiere Repository...")
            initialize_git_repo(current_dir)
            return
        
        # Hole Remote-Informationen
        subprocess.run(
            ["git", "fetch", "origin"],
            cwd=current_dir,
            capture_output=True,
            check=True
        )
        
        # Prüfe auf Unterschiede zwischen local und remote
        result = subprocess.run(
            ["git", "status", "-uno"],
            cwd=current_dir,
            capture_output=True,
            text=True
        )
        
        if "behind" in result.stdout.lower():
            print("📥 Neuere Version verfügbar! Starte Update...")
            pull_updates(current_dir)
        else:
            print("✓ Schon auf neuester Version")
            
    except Exception as e:
        print(f"❌ Fehler beim Update-Check: {e}")

def initialize_git_repo(directory):
    """Initialisiert ein Git-Repository"""
    try:
        subprocess.run(
            ["git", "init"],
            cwd=directory,
            check=True,
            capture_output=True
        )
        
        subprocess.run(
            ["git", "remote", "add", "origin", REPO_URL],
            cwd=directory,
            check=True,
            capture_output=True
        )
        
        subprocess.run(
            ["git", "fetch", "origin"],
            cwd=directory,
            check=True,
            capture_output=True
        )
        
        subprocess.run(
            ["git", "checkout", "-b", "main", "origin/main"],
            cwd=directory,
            capture_output=True
        )
        
        print("✓ Git-Repository initialisiert")
        
    except Exception as e:
        print(f"❌ Fehler beim Initialisieren: {e}")

def pull_updates(directory):
    """Pullt die neuesten Updates"""
    try:
        result = subprocess.run(
            ["git", "pull", "origin", "main"],
            cwd=directory,
            capture_output=True,
            text=True
        )
        
        if result.returncode == 0:
            print("✓ Updates erfolgreich installiert!")
            return True
        else:
            print(f"⚠ Pull-Fehler: {result.stderr}")
            return False
            
    except Exception as e:
        print(f"❌ Fehler beim Pullen: {e}")
        return False


def get_resource_path(relative_path):
    """ Liefert den richtigen Pfad, egal ob als .py oder .exe gestartet """
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)


def is_pid_running(pid):
    if pid <= 0:
        return False
    try:
        os.kill(pid, 0)
    except OSError as e:
        if e.errno == errno.ESRCH:
            return False
        if e.errno == errno.EPERM:
            return True
        return False
    except Exception:
        return False
    return True

# --- KONFIGURATION ---
ENTWICKLER_MODUS = False 
# Wir speichern Presets im Dokumente-Ordner oder lokal, damit sie beschreibbar bleiben
PRESET_DIR = "presets_design"
LAYOUT_DIR = "presets_layout"

for d in [PRESET_DIR, LAYOUT_DIR]:
    if not os.path.exists(d): 
        os.makedirs(d)

EXCEL_PATH_FILE = "excel_path.txt"
SETTINGS_FILE = "settings.json"

def load_settings():
    """Lädt Einstellungen, z. B. zuletzt ausgewähltes Design."""
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return {}


def save_settings(settings):
    """Speichert Einstellungen wie zuletzt gewähltes Design."""
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(settings, f)
    except:
        pass


def load_excel_path():
    """Lädt den gespeicherten Excel-Pfad aus der Datei"""
    if os.path.exists(EXCEL_PATH_FILE):
        try:
            with open(EXCEL_PATH_FILE, "r", encoding="utf-8") as f:
                path = f.read().strip()
                if path and os.path.exists(path):
                    return path
        except:
            pass
    return ""

def save_excel_path(path):
    """Speichert den Excel-Pfad in eine Datei"""
    try:
        with open(EXCEL_PATH_FILE, "w", encoding="utf-8") as f:
            f.write(path)
    except:
        pass

RAUMZELLEN = {
    "3 EDV": 9, "5 Besprechung": 16, "103/104 Schulung": 23,
    "106 Besprechung": 30, "110 Schulung": 37, "112 EDV": 44,
    "203 Schulung": 51, "205 Schulung": 58, "210 Besprechung": 65,
    "305 Mitarbeiterbüro": 72, "214 Schulung": 79, "303 Schulung": 86,
    "312 Schulung": 93
}

HOURS = ["08:00-09:30", "09:45-11:15", "11:30-13:00", "13:30-15:00", "15:15-16:30"]

def sortiere_etagen(raeume_dict):
    etagen = {"Erdgeschoss": [], "1. Obergeschoss": [], "2. Obergeschoss": [], "3. Obergeschoss": []}
    for raum in raeume_dict.keys():
        first_part = raum.split()[0].split('/')[0]
        try:
            num = int(''.join(filter(str.isdigit, first_part)))
            if num < 100: etagen["Erdgeschoss"].append((num, raum))
            elif num < 200: etagen["1. Obergeschoss"].append((num, raum))
            elif num < 300: etagen["2. Obergeschoss"].append((num, raum))
            else: etagen["3. Obergeschoss"].append((num, raum))
        except: continue
    ablauf = []
    for name in ["Erdgeschoss", "1. Obergeschoss", "2. Obergeschoss", "3. Obergeschoss"]:
        if etagen[name]:
            etagen[name].sort()
            ablauf.append({"typ": "tabelle", "inhalt": [r[1] for r in etagen[name]], "titel": name})
    return ablauf

class AdminInterface:
    def __init__(self, root):
        self.root = root
        self.root.title("Reinit Display Manager Ultimate")
        self.root.geometry("1150x800")
        self.root.configure(bg="#f0f0f0")
        
        self.config = {
            "font_family": "Arial", "font_size": 40, "tab_font_size": 10,
            "fg_color": "#ffffff", "bg_color": "#000000",
            "tab_header_bg": "#404040", "tab_cell_bg": "#ffffff", "tab_highlight": "#FFAA00"
        }
        
        self.settings = load_settings()
        self.folien_liste = []
        self.previews = {}

        self.main_frame = tk.Frame(root, bg="#f0f0f0")
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # --- LINKS: FOLIEN-VERWALTUNG ---
        list_container = tk.LabelFrame(self.main_frame, text=" Folien Ablauf & Layouts ", padx=10, pady=10)
        list_container.pack(side="left", fill="both", expand=True)
        
        lay_p_frame = tk.Frame(list_container)
        lay_p_frame.pack(fill="x", pady=(0, 10))
        self.layout_var = tk.StringVar()
        self.layout_menu = ttk.Combobox(lay_p_frame, textvariable=self.layout_var, state="readonly")
        self.layout_menu.pack(side="left", fill="x", expand=True)
        tk.Button(lay_p_frame, text="Laden", command=self.load_layout).pack(side="left", padx=2)
        tk.Button(lay_p_frame, text="Speichern", command=self.save_layout, bg="#9b59b6", fg="white").pack(side="left", padx=2)
        
        list_inner = tk.Frame(list_container)
        list_inner.pack(fill="both", expand=True)
        self.listbox = tk.Listbox(list_inner, font=("Consolas", 10))
        self.listbox.pack(side="left", fill="both", expand=True)

        arrow_frame = tk.Frame(list_inner)
        arrow_frame.pack(side="right", fill="y", padx=5)
        tk.Button(arrow_frame, text="▲", command=lambda: self.move_folie(-1)).pack(pady=5)
        tk.Button(arrow_frame, text="▼", command=lambda: self.move_folie(1)).pack(pady=5)
        
        btn_box = tk.Frame(list_container)
        btn_box.pack(fill="x", pady=10)
        tk.Button(btn_box, text="+ Text", command=self.add_text).pack(side="left", expand=True, padx=2)
        tk.Button(btn_box, text="+ Bild", command=self.add_bild).pack(side="left", expand=True, padx=2)
        tk.Button(btn_box, text="Tabellen (Auto)", command=self.add_auto_tables).pack(side="left", expand=True, padx=2)
        tk.Button(btn_box, text="Entfernen", fg="red", command=self.remove_folie).pack(side="right", expand=True, padx=2)

        # --- RECHTS: DESIGN-PANEL ---
        self.right_panel = tk.Frame(self.main_frame, bg="#f0f0f0")
        self.right_panel.pack(side="right", fill="both", padx=(10, 0))

        design_preset_frame = tk.LabelFrame(self.right_panel, text=" Design-Presets (Farben) ", padx=15, pady=10)
        design_preset_frame.pack(fill="x", pady=(0, 10))
        
        self.preset_var = tk.StringVar()
        self.preset_menu = ttk.Combobox(design_preset_frame, textvariable=self.preset_var, state="readonly")
        self.preset_menu.pack(fill="x", pady=5)
        
        tk.Button(design_preset_frame, text="Design laden", command=self.load_preset).pack(fill="x", pady=2)
        tk.Button(design_preset_frame, text="Design speichern", command=self.save_preset, bg="#3498db", fg="white").pack(fill="x", pady=2)

        design_frame = tk.LabelFrame(self.right_panel, text=" Design-Details ", padx=15, pady=10)
        design_frame.pack(fill="both", expand=True)

        self.entries = {}
        for label, key in [("Schriftart", "font_family"), ("Größe Text", "font_size"), ("Größe Tabelle", "tab_font_size")]:
            tk.Label(design_frame, text=label).pack(anchor="w")
            ent = tk.Entry(design_frame)
            ent.pack(fill="x", pady=2)
            self.entries[key] = ent

        color_fields = [("Textfarbe", "fg_color"), ("Hintergrund", "bg_color"), ("Tab-Kopf", "tab_header_bg"), ("Zellen Leer", "tab_cell_bg"), ("Zellen Belegt", "tab_highlight")]
        for label, key in color_fields:
            row = tk.Frame(design_frame); row.pack(fill="x", pady=2)
            preview = tk.Canvas(row, width=20, height=20, highlightthickness=1, highlightbackground="black")
            preview.pack(side="left", padx=(0, 5)); self.previews[key] = preview
            tk.Button(row, text=label, command=lambda k=key: self.pick_color(k)).pack(side="left", fill="x", expand=True)

        self.apply_config_to_ui()
        self.refresh_menus()
        self.add_auto_tables()

        # --- EXCEL-PFAD EINGABE ---
        excel_frame = tk.LabelFrame(root, text=" Excel-Datei ", padx=10, pady=10, bg="#f0f0f0")
        excel_frame.pack(fill="x", padx=20, pady=(0, 15))
        tk.Label(excel_frame, text="Pfad zur Excel-Datei:", bg="#f0f0f0").pack(anchor="w")
        self.excel_path_var = tk.StringVar(value=load_excel_path())
        self.excel_path_entry = tk.Entry(excel_frame, textvariable=self.excel_path_var, font=("Arial", 10))
        self.excel_path_entry.pack(fill="x", pady=(5, 0))

        path_btn_frame = tk.Frame(excel_frame, bg="#f0f0f0")
        path_btn_frame.pack(fill="x", pady=(5, 0))
        tk.Button(path_btn_frame, text="Durchsuchen", command=self.browse_excel).pack(side="left", padx=(0, 5))
        tk.Button(path_btn_frame, text="Speichern", command=self.save_excel_path_input, bg="#3498db", fg="white").pack(side="left")

        # --- AUTOMATISCHER PRÄSENTATIONSSTART ---
        auto_frame = tk.LabelFrame(root, text=" Automatischer Start ", padx=10, pady=10, bg="#f0f0f0")
        auto_frame.pack(fill="x", padx=20, pady=(0, 15))
        self.auto_start_seconds = tk.StringVar(value="60")
        self.auto_countdown_var = tk.StringVar(value="Warten...")
        auto_row = tk.Frame(auto_frame, bg="#f0f0f0")
        auto_row.pack(fill="x")
        tk.Label(auto_row, text="Start in", bg="#f0f0f0").pack(side="left")
        self.auto_start_entry = tk.Entry(auto_row, textvariable=self.auto_start_seconds, width=6)
        self.auto_start_entry.pack(side="left", padx=(5, 0))
        tk.Label(auto_row, text="Sekunden", bg="#f0f0f0").pack(side="left", padx=(5, 0))
        tk.Label(auto_row, textvariable=self.auto_countdown_var, bg="#f0f0f0").pack(side="left", padx=(10, 0))
        tk.Button(auto_frame, text="Timer neustarten", command=self.schedule_auto_start).pack(anchor="e", pady=(5, 0))

        tk.Button(root, text="PRÄSENTATION STARTEN", bg="#27ae60", fg="white", font=("Arial", 12, "bold"), height=2, command=self.start_show).pack(fill="x", padx=20, pady=15)
        self.update_listbox()

        self.auto_start_id = None
        self.auto_start_seconds_remaining = None
        self.schedule_auto_start()

    def refresh_menus(self):
        self.preset_menu['values'] = [f.replace(".json", "") for f in os.listdir(PRESET_DIR) if f.endswith(".json")]
        self.layout_menu['values'] = [f.replace(".json", "") for f in os.listdir(LAYOUT_DIR) if f.endswith(".json")]
        last_preset = self.settings.get("last_preset")
        if last_preset and last_preset in self.preset_menu['values']:
            self.preset_var.set(last_preset)

    def save_layout(self):
        name = simpledialog.askstring("Layout speichern", "Name:")
        if name:
            with open(os.path.join(LAYOUT_DIR, f"{name}.json"), "w", encoding='utf-8') as f:
                json.dump(self.folien_liste, f)
            self.refresh_menus()

    def load_layout(self):
        name = self.layout_var.get()
        if name:
            with open(os.path.join(LAYOUT_DIR, f"{name}.json"), "r", encoding='utf-8') as f:
                self.folien_liste = json.load(f)
            self.update_listbox()
            messagebox.showinfo("Layout", f"Layout '{name}' geladen.")

    def save_preset(self):
        name = simpledialog.askstring("Design", "Name:")
        if name:
            self.config.update({k: self.entries[k].get() for k in ["font_family", "font_size", "tab_font_size"]})
            with open(os.path.join(PRESET_DIR, f"{name}.json"), "w", encoding='utf-8') as f: json.dump(self.config, f)
            self.settings["last_preset"] = name
            save_settings(self.settings)
            self.refresh_menus()

    def load_preset(self):
        name = self.preset_var.get()
        if name:
            with open(os.path.join(PRESET_DIR, f"{name}.json"), "r", encoding='utf-8') as f: self.config = json.load(f)
            self.apply_config_to_ui()
            self.settings["last_preset"] = name
            save_settings(self.settings)

    def apply_config_to_ui(self):
        for k in ["font_family", "font_size", "tab_font_size"]:
            self.entries[k].delete(0, tk.END); self.entries[k].insert(0, str(self.config[k]))
        for k, v in self.previews.items(): v.configure(bg=self.config[k])
        # Apply overall background color to the admin UI
        try:
            self.root.configure(bg=self.config.get("bg_color", "#f0f0f0"))
            self.main_frame.configure(bg=self.config.get("bg_color", "#f0f0f0"))
            self.right_panel.configure(bg=self.config.get("bg_color", "#f0f0f0"))
        except Exception:
            pass

    def cancel_auto_start(self):
        if getattr(self, "auto_start_id", None):
            try:
                self.root.after_cancel(self.auto_start_id)
            except Exception:
                pass
            self.auto_start_id = None
        self.auto_start_seconds_remaining = None
        if hasattr(self, "auto_countdown_var"):
            self.auto_countdown_var.set("Abgebrochen")

    def schedule_auto_start(self):
        self.cancel_auto_start()
        try:
            seconds = int(self.auto_start_seconds.get())
        except Exception:
            self.auto_countdown_var.set("Ungültig")
            return
        if seconds <= 0:
            self.auto_countdown_var.set("Nicht gestartet")
            return
        self.auto_start_seconds_remaining = seconds
        self.auto_countdown_var.set(f"Startet in {seconds} s")
        self._auto_tick()

    def _auto_tick(self):
        if self.auto_start_seconds_remaining is None:
            return
        if self.auto_start_seconds_remaining <= 0:
            self.auto_countdown_var.set("Starte Präsentation...")
            self.start_show()
            return
        self.auto_countdown_var.set(f"Startet in {self.auto_start_seconds_remaining} s")
        self.auto_start_seconds_remaining -= 1
        self.auto_start_id = self.root.after(1000, self._auto_tick)


    def pick_color(self, key):
        c = colorchooser.askcolor(color=self.config[key])[1]
        if c:
            self.config[key] = c
            self.previews[key].configure(bg=c)
            self.apply_config_to_ui()

    def update_listbox(self):
        self.listbox.delete(0, tk.END)
        for i, f in enumerate(self.folien_liste):
            content = f['inhalt'].replace('\n', ' ')[:25] if f['typ'] == 'text' else (os.path.basename(f['inhalt']) if f['typ'] == 'bild' else f.get('titel', 'Tabelle'))
            self.listbox.insert(tk.END, f" {i+1:02} | [{f['typ'].upper():<7}] {content}")

    def move_folie(self, direction):
        sel = self.listbox.curselection()
        if sel:
            idx = sel[0]; new = idx + direction
            if 0 <= new < len(self.folien_liste):
                self.folien_liste[idx], self.folien_liste[new] = self.folien_liste[new], self.folien_liste[idx]
                self.update_listbox(); self.listbox.select_set(new)

    def add_text(self):
        t = simpledialog.askstring("Text", "Inhalt:"); 
        if t: self.folien_liste.append({"typ": "text", "inhalt": t}); self.update_listbox()

    def add_bild(self):
        p = filedialog.askopenfilename(filetypes=[("Bilder", "*.jpg *.png *.jpeg *.bmp")]); 
        if p: self.folien_liste.append({"typ": "bild", "inhalt": p}); self.update_listbox()

    def add_auto_tables(self):
        self.folien_liste.extend(sortiere_etagen(RAUMZELLEN))
        self.update_listbox()

    def remove_folie(self):
        sel = self.listbox.curselection()
        if sel: self.folien_liste.pop(sel[0]); self.update_listbox()

    def browse_excel(self):
        """Öffnet einen Dialog zum Auswählen der Excel-Datei."""
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")])
        if path:
            self.excel_path_var.set(path)
            save_excel_path(path)

    def save_excel_path_input(self):
        """Speichert den eingegebenen Excel-Pfad in einer Datei."""
        path = self.excel_path_var.get()
        if path and os.path.exists(path):
            save_excel_path(path)
            messagebox.showinfo("Erfolg", f"Excel-Pfad gespeichert:\n{path}")
        else:
            messagebox.showerror("Fehler", "Ungültiger Pfad oder Datei existiert nicht.")

    def start_show(self):
        self.cancel_auto_start()
        path = self.excel_path_var.get()
        if not path:
            messagebox.showerror("Fehler", "Bitte wählen Sie zuerst eine Excel-Datei aus.")
            return
        if not os.path.exists(path):
            messagebox.showerror("Fehler", f"Excel-Datei nicht gefunden:\n{path}")
            return
        
        # Prüfen, ob bereits eine Präsentation läuft
        presentation_lock = get_resource_path("presentation.lock")
        if os.path.exists(presentation_lock):
            try:
                with open(presentation_lock, "r", encoding="utf-8") as f:
                    pid = int(f.read().strip() or "0")
                if pid and is_pid_running(pid):
                    messagebox.showerror("Fehler", "Eine Präsentation läuft bereits.")
                    return
            except Exception:
                pass
            try:
                os.remove(presentation_lock)
            except Exception:
                pass
        
        # Lock erstellen
        try:
            with open(presentation_lock, "w", encoding="utf-8") as f:
                f.write(str(os.getpid()))
        except:
            messagebox.showerror("Fehler", "Konnte Lock-Datei nicht erstellen.")
            return
        
        self.config.update({k: self.entries[k].get() for k in ["font_family", "font_size", "tab_font_size"]})
        MasterPrasentation(tk.Toplevel(self.root), path, self.folien_liste, self.config, ENTWICKLER_MODUS)

class MasterPrasentation:
    def __init__(self, root, excel_path, ablauf, config, dev_mode):
        self.root = root; self.config = config; self.root.configure(background=config["bg_color"])
        self.root.attributes('-fullscreen', True)
        self.daten = self.lade_excel_daten(excel_path); self.ablauf = ablauf; self.index = 0
        self.main_frame = tk.Frame(self.root, bg=config["bg_color"]); self.main_frame.pack(expand=True, fill='both')
        self.footer = tk.Label(self.root, text="", font=("Arial", 16), bg=config["bg_color"], fg=config["fg_color"])
        self.footer.pack(side="bottom", pady=20)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.root.bind('<Escape>', self.on_close)
        self.update_clock(); self.naechste_folie()

    def on_close(self, event=None):
        presentation_lock = get_resource_path("presentation.lock")
        if os.path.exists(presentation_lock):
            try:
                os.remove(presentation_lock)
            except:
                pass
        self.root.destroy()

    def update_clock(self):
        self.footer.config(text=datetime.now().strftime("%d.%m.%Y | %H:%M:%S Uhr"))
        self.root.after(1000, self.update_clock)

    def lade_excel_daten(self, path):
        wb = load_workbook(path, data_only=True); ws = wb["Uebersicht"]
        return {r: [ws["F" + str(v + i)].value or "" for i in range(5)] for r, v in RAUMZELLEN.items()}

    def zeige_tabelle(self, raum_liste, titel):
        self.footer.pack(side="bottom", pady=20)
        tk.Label(self.main_frame, text=titel.upper(), font=(self.config["font_family"], 30, "bold"), bg=self.config["bg_color"], fg=self.config["fg_color"]).pack(pady=20)
        f = tk.Frame(self.main_frame, bg=self.config["bg_color"]); f.pack(expand=True, fill="both", padx=20)
        tf = (self.config["font_family"], int(self.config["tab_font_size"]))
        tk.Label(f, text="Zeit", font=tf, bg=self.config["tab_header_bg"], fg="white").grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        for c, r in enumerate(raum_liste, 1):
            tk.Label(f, text=r, font=tf, bg=self.config["tab_header_bg"], fg="white", wraplength=120).grid(row=0, column=c, sticky="nsew", padx=1, pady=1)
        for r_idx, zeit in enumerate(HOURS, 1):
            tk.Label(f, text=zeit, font=tf, bg=self.config["tab_cell_bg"]).grid(row=r_idx, column=0, sticky="nsew", padx=1, pady=1)
            for c_idx, raum in enumerate(raum_liste, 1):
                val = self.daten[raum][r_idx-1]
                tk.Label(f, text=val, font=tf, bg=self.config["tab_highlight"] if val else self.config["tab_cell_bg"], height=4, wraplength=150).grid(row=r_idx, column=c_idx, sticky="nsew", padx=1, pady=1)
        for i in range(len(raum_liste)+1): f.grid_columnconfigure(i, weight=1)

    def zeige_bild(self, pfad):
        if not os.path.exists(pfad):
            tk.Label(self.main_frame, text=f"BILD NICHT GEFUNDEN:\n{pfad}", fg="red", bg="black").pack(expand=True)
            return
        self.footer.pack_forget()
        img = Image.open(pfad); self.root.update()
        img = img.resize((self.root.winfo_width(), self.root.winfo_height()), Image.Resampling.LANCZOS)
        self.photo = ImageTk.PhotoImage(img); tk.Label(self.main_frame, image=self.photo, bg="black").pack()

    def naechste_folie(self):
        for w in self.main_frame.winfo_children(): w.destroy()
        if not self.ablauf: return
        f = self.ablauf[self.index]
        try:
            if f["typ"] == "text": 
                self.footer.pack(side="bottom", pady=20)
                tk.Label(self.main_frame, text=f["inhalt"], font=(self.config["font_family"], int(self.config["font_size"])), bg=self.config["bg_color"], fg=self.config["fg_color"], justify="center").pack(expand=True)
            elif f["typ"] == "bild": self.zeige_bild(f["inhalt"])
            else: self.zeige_tabelle(f["inhalt"], f["titel"])
        except Exception as e:
            tk.Label(self.main_frame, text=f"Fehler in Folie: {e}", fg="red").pack()
            
        self.index = (self.index + 1) % len(self.ablauf)
        self.root.after(8000, self.naechste_folie)

# --- STARTFUNKTION ---
def start_app():
    # Prüfe und hole Updates
    check_and_pull_updates()
    
    # Git Repository Check: Stellt sicher, dass das lokale Repo nicht vom GitHub Link ist
    project_dir = os.path.dirname(__file__)
    git_dir = os.path.join(project_dir, ".git")
    expected_github_url = "https://github.com/Lole2k05/Reinit-Raumplaner.git"
    
    if os.path.isdir(git_dir):
        try:
            # Prüfe, welches Remote-Repository eingebunden ist
            result = subprocess.run(["git", "remote", "get-url", "origin"], 
                                    capture_output=True, text=True, cwd=project_dir)
            current_remote = result.stdout.strip() if result.returncode == 0 else ""
            
            if current_remote == expected_github_url:
                print("WARNUNG: Dieses Repository ist mit dem GitHub Link verknüpft.")
                print(f"Remote URL: {current_remote}")
                print("Git pull wird NICHT ausgeführt, um lokale Änderungen zu schützen.")
            else:
                print(f"Repository Remote: {current_remote}")
        except Exception as e:
            print("Git Remote Check Fehler:", e)
    else:
        print("Kein Git-Repository im Projektordner. Git Check wird übersprungen.")
    
    lock_file = get_resource_path("app.lock")
    if os.path.exists(lock_file):
        try:
            with open(lock_file, "r", encoding="utf-8") as f:
                pid = int(f.read().strip() or "0")
            if pid and is_pid_running(pid):
                print("App läuft bereits. Mehrere Instanzen sind nicht erlaubt.")
                sys.exit(0)
        except Exception:
            pass
        try:
            os.remove(lock_file)
        except Exception:
            pass
    
    main_root = tk.Tk()
    AdminInterface(main_root)
    
    def cleanup():
        if os.path.exists(lock_file):
            os.remove(lock_file)
        main_root.destroy()
    
    main_root.protocol("WM_DELETE_WINDOW", cleanup)
    
    # Lock-File erstellen
    with open(lock_file, 'w') as f:
        f.write(str(os.getpid()))
    
    main_root.mainloop()

# Verhindert, dass beim Importieren durch den Launcher direkt etwas passiert
if __name__ == "__main__":
    start_app()